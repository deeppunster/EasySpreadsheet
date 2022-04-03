"""
BuildSimpleSpreadsheet.py - Build a simple spreadsheet from scratch.
"""
from pathlib import Path
from secrets import choice

from openpyxl import Workbook
from openpyxl.chart import (
    AreaChart3D,
    Reference,
    Series,
)

from Constants import (
    Area_3D_Chart_Info,
    RefData,
    SPREADSHEET_NAME_3,
    nbr_tests,
    test_data,
)


class ManageSpreadsheet:
    """
    Provide the functions for managing a spreadsheet.
    """

    def __init__(self):
        self.workbook = None
        self.active_sheet = None
        return

    def create_spreadsheet(self):
        """
        Create an active_sheet.

        :return:
        """
        self.workbook = Workbook()
        self.active_sheet = self.workbook.active
        return

    def create_new_tab(self, sheet_name):
        """
        Create another spreadsheet, make it active, and rename it.

        :param: sheet_name: name for the new sheet
        :return:
        """
        # count how many tabs are in the workbook already
        tab_count = len(self.workbook.sheetnames)
        self.active_sheet = self.workbook.create_sheet(
            title=sheet_name, index=tab_count
        )
        return

    def rename_current_spreadsheet(self, new_name: str):
        """
        Rename the label for the current sheet.

        :param new_name:
        :return:
        """
        self.active_sheet.title = new_name
        return

    def add_cell(self, cell_location: str, value):
        """
        Add a value at an absolute cell location.

        :param cell_location: absolute cell reference e.g. 'A1' or 'GG297'
        :param value: string, number, or formula
        :return:
        """
        self.active_sheet[cell_location] = value
        return

    def add_rows(self, data_group: list[list]):
        """
        Add one or more rows of data at the current location.

        :param: data_group: a list (rows) of lists (columns) of data
        :return:
        """
        for data_row in data_group:
            self.active_sheet.append(data_row)
        return

    def add_3d_area_chart(self, chart_info: Area_3D_Chart_Info):
        """
        Add an area 3D chart to the current sheet.

        :param chart_info:
        :return:
        """
        chart = AreaChart3D()
        chart.title = chart_info.title
        chart.style = chart_info.style
        chart.x_axis.title = chart_info.x_axis_title
        chart.y_axis.title = chart_info.y_axis_title
        chart.legend = chart_info.legend
        category_refs = chart_info.category_info
        data_refs = chart_info.data_info
        cat_ref = Reference(
            self.active_sheet,
            min_col=category_refs.min_col,
            min_row=category_refs.min_row,
            max_row=category_refs.max_row,
        )
        data_ref = Reference(
            self.active_sheet,
            min_col=data_refs.min_col,
            min_row=data_refs.min_row,
            max_col=data_refs.max_col,
            max_row=data_refs.max_row,
        )
        chart.add_data(data_ref, titles_from_data=data_refs.titles_included)
        chart.set_categories(cat_ref)
        chart_loc = 'A' + str(nbr_tests + 3)
        self.active_sheet.add_chart(chart, chart_loc)
        return

    def save_spreadsheet(self, filepath: Path, filename: str):
        """
        Save the spreadsheet to a specified location.

        :param filepath: Fully qualified path to the directory
        :param filename: file name for the spreadsheet
        :return:
        """
        file_loc: Path = filepath
        file_name: str = filename
        full_path = file_loc / file_name
        self.workbook.save(full_path)
        return


if __name__ == '__main__':
    my_work_area = Path.cwd()
    my_sheet = ManageSpreadsheet()
    my_sheet.create_spreadsheet()
    my_sheet.rename_current_spreadsheet('Python')
    my_sheet.add_cell('A1', 'Hi from Python!')

    # add some random numbers to a column
    column = 'c'
    my_sheet.add_cell(column + '1', 'Random Numbers')
    for row in range(2, 12):
        ref = column + str(row)
        contents = choice(range(1000))
        my_sheet.add_cell(ref, contents)

    # add a new sheet, add data, and make a chart
    my_sheet.create_new_tab('Test Data Chart')
    my_sheet.add_rows(test_data)

    category_info = RefData(
        min_col=1,
        max_col=0,  # not used
        min_row=1,
        max_row=nbr_tests,
        titles_included=False,
    )
    data_info = RefData(
        min_col=2,
        max_col=3,
        min_row=1,
        max_row=nbr_tests,
        titles_included=True,
    )
    chart_info = Area_3D_Chart_Info(
        title='Area Chart',
        style=13,
        x_axis_title='Test',
        y_axis_title='Percentage',
        legend=None,
        category_info=category_info,
        data_info=data_info,
    )
    my_sheet.add_3d_area_chart(chart_info)
    my_sheet.save_spreadsheet(my_work_area, SPREADSHEET_NAME_3)
